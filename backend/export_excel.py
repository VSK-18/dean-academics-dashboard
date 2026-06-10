import sys
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

def export_excel(json_data_path, template_path, output_path):
    # Load transactions
    with open(json_data_path, 'r', encoding='utf-8') as f:
        proposals = json.load(f)

    # Sort proposals by date
    proposals.sort(key=lambda x: x.get('date', ''))

    # Load Excel Workbook
    wb = openpyxl.load_workbook(template_path)

    # 1. Populate AAF-MONITOR and EQA-Monitor
    for cat in ['AAF', 'EQA']:
        sheet_name = 'AAF-MONITOR' if cat == 'AAF' else 'EQA-Monitor'
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        
        # Find insertion row (we look for the row containing "Total Proposed" in Column F)
        total_row_idx = None
        for r in range(20, 100):
            val = ws.cell(row=r, column=6).value
            if val and "Total Proposed" in str(val):
                total_row_idx = r
                break
        
        if not total_row_idx:
            # Fallback if not found
            total_row_idx = 33
            
        # Get transactions for this category
        cat_proposals = [p for p in proposals if p.get('category') == cat]
        
        # We start writing at row 23.
        # How many rows do we need?
        num_records = len(cat_proposals)
        
        # Clear existing demo rows first (rows 23 to total_row_idx-1)
        # Note: B is Ref, C is Date, D is Description, F is Dept, G is Sub-head, H is Proposed, K is Actual, L is Month, M is Remarks, N is Balance
        for r in range(23, total_row_idx):
            ws.cell(row=r, column=2).value = None
            ws.cell(row=r, column=3).value = None
            ws.cell(row=r, column=4).value = None
            ws.cell(row=r, column=6).value = None
            ws.cell(row=r, column=7).value = None
            ws.cell(row=r, column=8).value = None
            ws.cell(row=r, column=11).value = None
            ws.cell(row=r, column=12).value = None
            ws.cell(row=r, column=13).value = None
            # Do not touch Column N which has formulas: =N(r-1) - K(r)
            
        # If the number of records is larger than the available rows, insert rows
        available_rows = (total_row_idx - 1) - 23 + 1
        if num_records > available_rows:
            diff = num_records - available_rows
            ws.insert_rows(total_row_idx - 1, amount=diff)
            # Copy formulas and formatting for the inserted rows in Column N
            for i in range(diff):
                curr_row = total_row_idx - 1 + i
                ws.cell(row=curr_row, column=14).value = f"=N{curr_row-1}-K{curr_row}"
            # Update total row index
            total_row_idx += diff
            
        # Now fill in the data
        for idx, p in enumerate(cat_proposals):
            row_num = 23 + idx
            
            # Format date to DD.MM.YYYY
            dt_str = p.get('date', '')
            try:
                dt_obj = datetime.strptime(dt_str, '%Y-%m-%d')
                formatted_date = dt_obj.strftime('%d.%m.%Y')
            except Exception:
                formatted_date = dt_str

            ws.cell(row=row_num, column=2).value = p.get('id', '')
            ws.cell(row=row_num, column=3).value = formatted_date
            ws.cell(row=row_num, column=4).value = p.get('comp', '')
            ws.cell(row=row_num, column=6).value = p.get('dept', '')
            ws.cell(row=row_num, column=7).value = p.get('subHead', '')
            ws.cell(row=row_num, column=8).value = float(p.get('amountProposed', 0))
            ws.cell(row=row_num, column=11).value = float(p.get('actualExpenditure', 0))
            ws.cell(row=row_num, column=12).value = p.get('month', '')
            ws.cell(row=row_num, column=13).value = p.get('remarks', '')
            
            # Ensure balance formula exists in Column N
            if row_num > 23:
                ws.cell(row=row_num, column=14).value = f"=N{row_num-1}-K{row_num}"
            else:
                ws.cell(row=row_num, column=14).value = f"=6200000-K23" if cat == 'AAF' else f"=10000000-K23"

    # 2. Populate Monthly Sheet
    if 'Monthly' in wb.sheetnames:
        ws_monthly = wb['Monthly']
        
        # Clear existing contents starting at row 2 (index 2, since row 1 has headers)
        # Note: Columns A-G for AAF, H-N for EQA
        max_row = ws_monthly.max_row
        if max_row >= 2:
            for r in range(2, max_row + 1):
                for col in range(1, 15):
                    ws_monthly.cell(row=r, column=col).value = None
                    
        # Write fresh data
        aaf_proposals = [p for p in proposals if p.get('category') == 'AAF']
        eqa_proposals = [p for p in proposals if p.get('category') == 'EQA']
        
        total_rows = max(len(aaf_proposals), len(eqa_proposals))
        
        for idx in range(total_rows):
            row_num = 2 + idx
            
            # Write AAF
            if idx < len(aaf_proposals):
                p = aaf_proposals[idx]
                dt_str = p.get('date', '')
                try:
                    dt_obj = datetime.strptime(dt_str, '%Y-%m-%d')
                    formatted_date = dt_obj.strftime('%Y-%m-%d')
                except Exception:
                    formatted_date = dt_str
                    
                ws_monthly.cell(row=row_num, column=1).value = formatted_date
                ws_monthly.cell(row=row_num, column=3).value = p.get('id', '')
                ws_monthly.cell(row=row_num, column=4).value = p.get('comp', '')
                ws_monthly.cell(row=row_num, column=5).value = p.get('dept', '')
                ws_monthly.cell(row=row_num, column=6).value = float(p.get('actualExpenditure', 0))
                
            # Write EQA
            if idx < len(eqa_proposals):
                p = eqa_proposals[idx]
                dt_str = p.get('date', '')
                try:
                    dt_obj = datetime.strptime(dt_str, '%Y-%m-%d')
                    formatted_date = dt_obj.strftime('%Y-%m-%d')
                except Exception:
                    formatted_date = dt_str
                    
                ws_monthly.cell(row=row_num, column=8).value = formatted_date
                ws_monthly.cell(row=row_num, column=10).value = p.get('id', '')
                ws_monthly.cell(row=row_num, column=11).value = p.get('comp', '')
                ws_monthly.cell(row=row_num, column=12).value = p.get('dept', '')
                ws_monthly.cell(row=row_num, column=13).value = float(p.get('actualExpenditure', 0))

    # Save output
    wb.save(output_path)
    print("Success")

if __name__ == '__main__':
    if len(sys.argv) < 4:
        print("Usage: python export_excel.py <json_path> <template_path> <output_path>")
        sys.exit(1)
    export_excel(sys.argv[1], sys.argv[2], sys.argv[3])
